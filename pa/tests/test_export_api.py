from io import BytesIO
from django.contrib.auth import get_user_model
from django.test import TestCase
from unittest.mock import patch
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from pa.models import Plan, Action, Profile
from pa.exporters import export_actions_excel


class ExportAPITests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.sa = User.objects.create_user("sa", password="pass")
        self.sa_profile = Profile.objects.create(user=self.sa, role=Profile.Role.SUPER_ADMIN)
        self.pilot = User.objects.create_user("pilot", password="pass")
        self.pilot_profile = Profile.objects.create(user=self.pilot, role=Profile.Role.PILOTE)
        self.plan1 = Plan.objects.create(nom="Plan1", excel_path="p1.xlsx", excel_sheet="plan d’action", header_row_index=11)
        self.plan2 = Plan.objects.create(nom="Plan2", excel_path="p2.xlsx", excel_sheet="plan d’action", header_row_index=11)
        self.pilot_profile.plans_autorises.add(self.plan1)
        Action.objects.create(
            act_id="ACT-0001",
            titre="A1",
            statut="En cours",
            priorite="High",
            plan=self.plan1,
            excel_fichier="p1.xlsx",
            excel_feuille="plan d’action",
            excel_row_index=1,
            j=5,
        )
        Action.objects.create(
            act_id="ACT-0002",
            titre="A2",
            statut="En cours",
            priorite="Low",
            plan=self.plan2,
            excel_fichier="p2.xlsx",
            excel_feuille="plan d’action",
            excel_row_index=1,
            j=5,
        )

    def test_excel_export_respects_roles_and_filters(self):
        from rest_framework.test import APIClient

        client = APIClient()
        client.force_authenticate(self.sa)
        resp = client.get("/api/export/excel")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("actions_", resp["Content-Disposition"])
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws[1][0].value, "ACT-ID")
        ids = {row[0].value for row in ws.iter_rows(min_row=2)}
        self.assertIn("ACT-0001", ids)
        self.assertIn("ACT-0002", ids)
        resp = client.get("/api/export/excel", {"plan": self.plan1.id})
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        ids = {row[0].value for row in ws.iter_rows(min_row=2)}
        self.assertEqual(ids, {"ACT-0001"})
        client.force_authenticate(self.pilot)
        resp = client.get("/api/export/excel")
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        ids = {row[0].value for row in ws.iter_rows(min_row=2)}
        self.assertEqual(ids, {"ACT-0001"})

    def test_pdf_export_respects_roles_and_filters(self):
        from rest_framework.test import APIClient

        client = APIClient()
        client.force_authenticate(self.sa)
        resp = client.get("/api/export/pdf")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        reader = PdfReader(BytesIO(resp.content))
        text = "".join(page.extract_text() or "" for page in reader.pages)
        self.assertIn("ACT-0001", text)
        self.assertIn("ACT-0002", text)
        resp = client.get("/api/export/pdf", {"plan": self.plan1.id})
        reader = PdfReader(BytesIO(resp.content))
        text = "".join(page.extract_text() or "" for page in reader.pages)
        self.assertIn("ACT-0001", text)
        self.assertNotIn("ACT-0002", text)
        client.force_authenticate(self.pilot)
        resp = client.get("/api/export/pdf")
        reader = PdfReader(BytesIO(resp.content))
        text = "".join(page.extract_text() or "" for page in reader.pages)
        self.assertIn("ACT-0001", text)
        self.assertNotIn("ACT-0002", text)

    def test_excel_export_large_queryset_chunks(self):
        qs = Action.objects.all()
        with patch.object(qs, "count", return_value=50001), patch.object(qs, "iterator", wraps=qs.iterator) as mock_iter:
            export_actions_excel(qs)
            mock_iter.assert_called_with(chunk_size=5000)
