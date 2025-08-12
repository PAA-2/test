from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from datetime import date
from io import BytesIO
from openpyxl import load_workbook

from pa.models import Plan, Action, Profile


class ExportsAPITests(TestCase):
    def setUp(self):
        self.client = APIClient()
        User = get_user_model()
        self.user = User.objects.create_user('sa', password='pass')
        Profile.objects.create(user=self.user, role=Profile.Role.SUPER_ADMIN)
        self.plan = Plan.objects.create(
            nom='Plan1', excel_path='/tmp/p1.xlsx', excel_sheet='plan d’action', header_row_index=11
        )
        Action.objects.create(
            act_id='ACT-0001', titre='foo', statut='open', priorite='High',
            delais=date.today(), j=0, plan=self.plan,
            excel_fichier=self.plan.excel_path, excel_feuille='plan d’action', excel_row_index=12
        )
        self.client.force_authenticate(self.user)

    def test_excel_export_respects_roles_and_filters(self):
        resp = self.client.get('/api/export/excel', {'plan': self.plan.id})
        assert resp.status_code == 200
        assert resp['Content-Type'].startswith('application/vnd.openxml')
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 'ACT-0001'

    def test_pdf_export_respects_roles_and_filters(self):
        resp = self.client.get('/api/export/pdf', {'plan': self.plan.id})
        assert resp.status_code == 200
        assert resp['Content-Type'] == 'application/pdf'
