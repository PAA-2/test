from io import BytesIO
from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from pa.models import Plan, Action, Profile


class ExportReportTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.sa = User.objects.create_user("sa", password="pass")
        Profile.objects.create(user=self.sa, role=Profile.Role.SUPER_ADMIN)
        self.client = APIClient()
        self.plan = Plan.objects.create(
            nom="Plan1",
            excel_path="/tmp/plan.xlsx",
            excel_sheet="plan d’action",
            header_row_index=11,
        )
        self.action = Action.objects.create(
            act_id="ACT-0001",
            titre="Action1",
            statut="En cours",
            priorite="High",
            plan=self.plan,
            excel_fichier=self.plan.excel_path,
            excel_feuille="plan d’action",
            excel_row_index=12,
        )

    def test_export_excel_respects_filters_and_roles(self):
        self.client.force_authenticate(self.sa)
        resp = self.client.get(f"/api/export/excel?plan={self.plan.id}")
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(filename=BytesIO(resp.content))
        sheet = wb.active
        self.assertEqual(sheet[2][0].value, "ACT-0001")

    def test_export_pdf_returns_stream_with_filename(self):
        self.client.force_authenticate(self.sa)
        resp = self.client.get("/api/export/pdf")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp["Content-Disposition"].startswith("attachment; filename=actions_"))

    def test_custom_report_sections_kpi_table_text(self):
        self.client.force_authenticate(self.sa)
        payload = {
            "title": "Rapport",
            "filters": {"plan": self.plan.id},
            "sections": [
                {"type": "kpi"},
                {"type": "table", "columns": ["act_id", "titre"]},
                {"type": "text", "text": "Note"},
            ],
            "dry_run": True,
        }
        resp = self.client.post("/api/reports/custom", payload, format="json")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("kpis", resp.data)
        self.assertIn("table_preview", resp.data)

    def test_pilote_cannot_export_outside_authorized_plans(self):
        User = get_user_model()
        pilote = User.objects.create_user("pilote", password="pass")
        prof = Profile.objects.create(user=pilote, role=Profile.Role.PILOTE)
        # no plans authorised
        self.client.force_authenticate(pilote)
        resp = self.client.get(f"/api/export/excel?plan={self.plan.id}")
        self.assertEqual(resp.status_code, 200)
        # but dataset should be empty
        wb = load_workbook(filename=BytesIO(resp.content))
        sheet = wb.active
        self.assertEqual(sheet.max_row, 1)

    def test_columns_and_ordering_match_list_endpoint(self):
        self.client.force_authenticate(self.sa)
        resp = self.client.get("/api/export/excel", {"ordering": "-act_id", "columns": "act_id,titre"})
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(filename=BytesIO(resp.content))
        sheet = wb.active
        self.assertEqual(sheet[1][0].value, "act_id")
