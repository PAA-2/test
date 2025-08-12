import logging
from datetime import datetime
from io import BytesIO

import pandas as pd
from django.db.models import Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from ..models import Action
from ..charts import _actions_for_user

logger = logging.getLogger(__name__)

DEFAULT_COLUMNS = [
    "act_id",
    "titre",
    "statut",
    "priorite",
    "J",
    "delais",
    "responsables",
    "plan",
    "date_creation",
    "date_realisation",
]

ORDERING_FIELDS = [
    "act_id",
    "titre",
    "priorite",
    "statut",
    "delais",
    "date_creation",
    "j",
]


def query_actions_for_export(user, params):
    qs = _actions_for_user(user).select_related("plan").prefetch_related("responsables")
    plan = params.get("plan")
    statut = params.get("statut")
    priorite = params.get("priorite")
    responsable = params.get("responsable")
    q = params.get("q")
    date_from = params.get("from")
    date_to = params.get("to")
    ordering = params.get("ordering")

    if plan:
        qs = qs.filter(plan_id=plan)
    if statut:
        qs = qs.filter(statut=statut)
    if priorite:
        qs = qs.filter(priorite=priorite)
    if responsable:
        qs = qs.filter(responsables__username=responsable)
    if q:
        qs = qs.filter(Q(titre__icontains=q) | Q(commentaire__icontains=q))
    if date_from:
        qs = qs.filter(date_creation__gte=date_from)
    if date_to:
        qs = qs.filter(date_creation__lte=date_to)
    if ordering:
        field = ordering.lstrip("-")
        if field not in ORDERING_FIELDS:
            from rest_framework.exceptions import ValidationError

            raise ValidationError({"ordering": "Invalid field"})
        qs = qs.order_by(ordering)
    return qs


def _rows_from_queryset(qs, columns):
    rows = []
    for a in qs:
        responsables = ", ".join(a.responsables.values_list("username", flat=True))
        row = {
            "act_id": a.act_id,
            "titre": a.titre,
            "statut": a.statut,
            "priorite": a.priorite,
            "J": a.j,
            "delais": a.delais.isoformat() if a.delais else "",
            "responsables": responsables,
            "plan": a.plan.nom if a.plan else "",
            "date_creation": a.date_creation.isoformat() if a.date_creation else "",
            "date_realisation": a.date_realisation.isoformat() if a.date_realisation else "",
        }
        rows.append([row.get(c, "") for c in columns])
    return rows


def to_excel(qs, columns=None):
    columns = columns or list(DEFAULT_COLUMNS)
    data = _rows_from_queryset(qs, columns)
    df = pd.DataFrame(data, columns=columns)
    stream = BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Actions")
        ws = writer.sheets["Actions"]
        ws.freeze_panes = "A2"
        from openpyxl.styles import Font, PatternFill

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    stream.seek(0)
    return stream


def to_pdf_table(qs, columns=None):
    columns = columns or list(DEFAULT_COLUMNS)
    data = _rows_from_queryset(qs, columns)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    table = Table([columns] + data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    doc.build([table])
    buffer.seek(0)
    return buffer
