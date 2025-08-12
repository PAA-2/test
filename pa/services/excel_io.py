import logging
import os
from typing import List, Dict

import pandas as pd
from django.conf import settings
from openpyxl import load_workbook

from pa.models import Plan, Action

logger = logging.getLogger(__name__)


def resolve_excel_path(path: str) -> str:
    if os.path.isabs(path):
        return path
    root = getattr(settings, "EXCEL_ROOT", None)
    if root:
        return os.path.join(root, path)
    return path


def read_plan(plan: Plan, limit: int = 50) -> List[Dict]:
    """Return a preview of plan actions from the source Excel file."""
    path = resolve_excel_path(plan.excel_path)
    df = pd.read_excel(
        path,
        sheet_name=plan.excel_sheet,
        header=plan.header_row_index - 1,
    )
    return df.head(limit).to_dict(orient="records")


def write_action(action: Action) -> Action:
    """Write the action back to its Excel source and reload it from disk."""
    file_path = resolve_excel_path(action.excel_fichier)
    wb = load_workbook(file_path)
    ws = wb[action.excel_feuille]
    header_row = action.plan.header_row_index
    headers = [cell.value for cell in ws[header_row]]
    col_map = {header: idx + 1 for idx, header in enumerate(headers) if header}
    row = action.excel_row_index

    if "act_id" in col_map:
        ws.cell(row=row, column=col_map["act_id"], value=action.act_id)
    if "titre" in col_map:
        ws.cell(row=row, column=col_map["titre"], value=action.titre)
    if "statut" in col_map:
        ws.cell(row=row, column=col_map["statut"], value=action.statut)
    if "priorite" in col_map:
        ws.cell(row=row, column=col_map["priorite"], value=action.priorite)

    wb.save(file_path)
    logger.info(
        "Action %s written to %s[%s] row %s",
        action.act_id,
        file_path,
        action.excel_feuille,
        action.excel_row_index,
    )

    df = pd.read_excel(
        file_path,
        sheet_name=action.excel_feuille,
        header=action.plan.header_row_index - 1,
    )
    row_data = df[df["act_id"] == action.act_id].iloc[0].to_dict()
    for field in ["titre", "statut", "priorite"]:
        if field in row_data:
            setattr(action, field, row_data[field])
    action.save()
    return action


def apply_update(act_id: str, strategy: str = "plan") -> int:
    """Apply updates for all occurrences of an action depending on strategy."""
    actions = Action.objects.filter(act_id=act_id)
    if not actions.exists():
        return 0
    if strategy == "all":
        to_update = actions
    elif strategy == "active":
        to_update = actions.exclude(statut__iexact="closed")
    else:  # default 'plan'
        to_update = [actions.first()]
    count = 0
    for act in to_update:
        write_action(act)
        count += 1
    return count
