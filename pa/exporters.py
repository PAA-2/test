from __future__ import annotations

import logging
from io import BytesIO
from typing import Dict, Any

import pandas as pd
from django.conf import settings
from django.db.models import Q, QuerySet
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .models import Action, Profile

logger = logging.getLogger("audit")


COLUMNS = [
    "ACT-ID",
    "Titre",
    "Statut",
    "Priorité",
    "Budget (DZD)",
    "Responsables",
    "Délais",
    "Date réalisation",
    "J",
    "P",
    "D",
    "C",
    "A",
    "Plan",
    "Fichier",
    "Feuille",
    "RowIndex",
]


def build_actions_queryset(user, params) -> QuerySet[Action]:
    qs = Action.objects.all()

    role = getattr(user.profile, "role", None)
    if role in (Profile.Role.PILOTE, Profile.Role.UTILISATEUR):
        allowed = user.profile.plans_autorises.all()
        qs = qs.filter(plan__in=allowed)

    # filters
    if plan := params.get("plan"):
        qs = qs.filter(plan_id=plan)
    if statut := params.get("statut"):
        qs = qs.filter(statut=statut)
    if priorite := params.get("priorite"):
        qs = qs.filter(priorite=priorite)
    if responsable := params.get("responsable"):
        qs = qs.filter(responsables__id=responsable)
    if q := params.get("q"):
        qs = qs.filter(Q(titre__icontains=q) | Q(commentaire__icontains=q))
    if from_date := params.get("from"):
        qs = qs.filter(date_creation__gte=from_date)
    if to_date := params.get("to"):
        qs = qs.filter(date_creation__lte=to_date)

    return qs.distinct()


def _action_to_row(action: Action) -> Dict[str, Any]:
    responsables = ",".join(action.responsables.values_list("username", flat=True))
    return {
        "ACT-ID": action.act_id,
        "Titre": action.titre,
        "Statut": action.statut,
        "Priorité": action.priorite,
        "Budget (DZD)": f"{action.budget_dzd:.2f}" if action.budget_dzd is not None else "",
        "Responsables": responsables,
        "Délais": action.delais.strftime("%Y-%m-%d") if action.delais else "",
        "Date réalisation": action.date_realisation.strftime("%Y-%m-%d") if action.date_realisation else "",
        "J": action.j if action.j is not None else "",
        "P": int(action.p),
        "D": int(action.d),
        "C": int(action.c),
        "A": int(action.a),
        "Plan": action.plan.nom,
        "Fichier": action.excel_fichier,
        "Feuille": action.excel_feuille,
        "RowIndex": action.excel_row_index,
    }


def export_actions_excel(qs: QuerySet[Action], filters: Dict[str, Any]) -> BytesIO:
    total = qs.count()
    output = BytesIO()
    if total <= 50_000:
        data = [_action_to_row(a) for a in qs]
        df = pd.DataFrame(data, columns=COLUMNS)
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Actions")
            ws = writer.sheets["Actions"]
            header_font = Font(bold=True)
            fill = PatternFill("solid", fgColor="DDDDDD")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = fill
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
            ws.freeze_panes = "A2"
    else:
        from openpyxl.cell import WriteOnlyCell

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Actions")
        header_font = Font(bold=True)
        fill = PatternFill("solid", fgColor="DDDDDD")
        header_cells = []
        for col in COLUMNS:
            cell = WriteOnlyCell(ws, value=col)
            cell.font = header_font
            cell.fill = fill
            header_cells.append(cell)
        ws.append(header_cells)
        max_widths = [len(col) for col in COLUMNS]
        batch_size = 5000
        for start in range(0, total, batch_size):
            for action in qs[start : start + batch_size]:
                row = list(_action_to_row(action).values())
                ws.append(row)
                max_widths = [
                    max(w, len(str(val)) if val is not None else 0)
                    for w, val in zip(max_widths, row)
                ]
        for i, width in enumerate(max_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width + 2
        ws.freeze_panes = "A2"
        wb.save(output)

    output.seek(0)
    logger.info(
        "EXPORT",
        extra={"payload": {"filters": dict(filters), "format": "excel", "count": total}},
    )
    return output


def export_actions_pdf(qs: QuerySet[Action], filters: Dict[str, Any]) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    logo_path = settings.BASE_DIR / "static" / "logo.png"
    if logo_path.exists():
        elements.append(Image(str(logo_path), width=50, height=50))
    elements.append(Paragraph("Export Actions", styles["Title"]))
    elements.append(Spacer(1, 12))

    cols = COLUMNS[:12] if len(COLUMNS) > 12 else COLUMNS
    data = [cols]
    for action in qs:
        row_dict = _action_to_row(action)
        row = [row_dict[c] for c in cols]
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    elements.append(table)

    def footer(canvas, doc):
        canvas.saveState()
        text = timezone.now().strftime("%Y-%m-%d")
        canvas.drawString(40, 20, text)
        summary = ", ".join(f"{k}={v}" for k, v in filters.items())
        canvas.drawRightString(doc.width + 100, 20, summary)
        canvas.restoreState()

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    logger.info(
        "EXPORT",
        extra={"payload": {"filters": dict(filters), "format": "pdf", "count": qs.count()}},
    )
    return buffer

